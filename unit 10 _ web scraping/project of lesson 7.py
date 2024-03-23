from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from pathlib import Path

excel_file_path = Path.home() / Path('Desktop', 'wikiPedia.xlsx')

driver = webdriver.Chrome()

try:
    driver.get('https://en.wikipedia.org/wiki/List_of_languages_by_number_of_native_speakers')

    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, 'table.wikitable'))
    )

    table = driver.find_element(By.CSS_SELECTOR, 'table.wikitable')
    rows = table.find_elements(By.TAG_NAME, 'tr')

    headers = [head.text.replace('\n', '') for head in rows[0].find_elements(By.TAG_NAME, 'th')]
    data = []

    for row in rows[1:]:
        cells = row.find_elements(By.TAG_NAME, 'td')
        row_data = [cell.text.strip() for cell in cells]
        data.append(row_data)

except Exception as e:

    print(f"An error occurred: {e}")


finally:

    driver.quit()


excel_file = Workbook()
excel_sheet = excel_file.active

for col_num, header in enumerate(headers, 1):
    excel_sheet.cell(row=1, column=col_num, value=header)

for row_num, row_data in enumerate(data, 2):
    for col_num, value in enumerate(row_data, 1):
        excel_sheet.cell(row=row_num, column=col_num, value=value)

excel_file.save(excel_file_path)